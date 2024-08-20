import time
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import os

# 엑셀 파일 생성
book = Workbook()
sheet = book.active
sheet.title = "Blog Data"
sheet.append(["Title", "Image"])

# 웹드라이버 설정
keyword = "AI"
opt = webdriver.ChromeOptions()
opt.add_experimental_option("detach", True)
browser = webdriver.Chrome(options=opt)

# 검색 결과 1~3페이지 본문 URL 수집
blog_url_list = []
for page_num in range(1, 4):
    browser.get(f"https://section.blog.naver.com/Search/Post.naver?pageNo={page_num}&rangeType=ALL&orderBy=sim&keyword={keyword}")
    time.sleep(2)
    items = browser.find_elements(By.CSS_SELECTOR, "a.desc_inner")
    for i in items:
        blog_url_list.append(i.get_attribute("href"))

# 이미지 저장 폴더 생성
image_folder = 'blog_images'
os.makedirs(image_folder, exist_ok=True)

num = 1
for blog_url in blog_url_list:
    blog_url_for_mobile = blog_url.replace("https://blog.naver", "https://m.blog.naver")
    browser.get(blog_url_for_mobile)
    time.sleep(2)
    
    # 제목 추출
    try:
        title = browser.find_element(By.CSS_SELECTOR, "div.se-component-content span").text
    except:
        title = "No Title"
    
    # 사진 URL 추출 및 저장
    try:
        image_element = browser.find_element(By.CSS_SELECTOR, "div.se-main-container img")
        image_url = image_element.get_attribute("src")
        img_path = f"{image_folder}/{num}.jpg"
        img_response = requests.get(image_url, stream=True)
        with open(img_path, 'wb') as img_file:
            img_file.write(img_response.content)
        
        # 이미지 크기 조절
        img = PILImage.open(img_path)
        img_resized = img.resize((int(img.width * 0.3), int(img.height * 0.3)))  # 크기 조절
        img_resized_path = f"{image_folder}/{num}_resized.jpg"
        img_resized.save(img_resized_path)
        img.close()
    except:
        image_url = "No Image"
        img_resized_path = None
    
    # 블로그 내용 추출
    try:
        content = browser.find_element(By.CSS_SELECTOR, "div.se-main-container").text
    except:
        content = "No Content"
    
    # 엑셀에 제목과 이미지 추가
    sheet.cell(row=num + 1, column=1).value = title
    if img_resized_path:
        excel_img = Image(img_resized_path)
        excel_img.width = 150  # 엑셀에서의 이미지 너비 (포인트)
        excel_img.height = 200  # 엑셀에서의 이미지 높이 (포인트)
        sheet.add_image(excel_img, f"B{num + 1}")
    
    # 블로그 내용 텍스트 파일에 저장
    with open('blog_contents.txt', 'a', encoding='utf-8') as text_file:
        text_file.write(f"URL: {blog_url_for_mobile}\n")
        text_file.write(f"Title: {title}\n")
        text_file.write(content.replace("\n", " ").strip())
        text_file.write("\n" + ("-" * 50) + "\n")
    
    print(f"제목 : {title}")
    print(f"이미지 URL : {image_url}")
    print(f"내용 : {content.replace('\n', ' ').strip()}")
    print("--------------------------------")
    
    num += 1

# 열 너비 조절
sheet.column_dimensions["A"].width = 35
sheet.column_dimensions["B"].width = 50

# 행 높이 조절
for row in sheet.iter_rows(min_row=2, max_row=num):
    sheet.row_dimensions[row[0].row].height = 200  # 충분히 큰 행 높이 설정
    
# 엑셀 파일 저장
book.save("blog_data.xlsx")

# 브라우저 닫기
browser.quit()
